import os, webbrowser, Pmw, logging, subprocess, pickle, shutil, sys
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Color, colors
import tkinter as tk
from tkinter import ttk
from tkinter.filedialog import askdirectory, askopenfilename
import tkinter.messagebox as msg_box
from xmldiff import main
from datetime import date, datetime, timedelta
import multiprocessing
import threading
# from multiprocessing import Queue
from queue import Queue


class ScrollableFrame(tk.Frame):

    def __init__(self, container, bg=None, *args, **kwargs):
        super().__init__(container, *args, **kwargs)
        self.canvas = tk.Canvas(self, bg="black")
        scrollbar = ttk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = tk.Frame(self.canvas)

        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all"))
            # lambda e: self.change_view()
        )
        self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=scrollbar.set)
        self.canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

    def change_view(self):
        self.canvas.config(scrollregion=self.canvas.bbox("all"))
        # self.canvas.yview_moveto(1)

    def move_to_end(self):
        self.canvas.yview_moveto(1)


def create_results_excel():
    global results_xlsx
    results_xlsx = os.path.join(report_folder, "xml_compare.xlsx")
    if not os.path.exists(results_xlsx):
        wb = Workbook()
        ws = wb.active
        ws.append(["Timestamp", "Old XML", "New XML", "Status"])
        for c in ws[1]:
            c.font = Font(bold=True)
        wb.save(results_xlsx)


def _set_reports_dir():
    global report_folder
    try:
        if dir_var['reports'].get():
            report_folder = os.path.join(dir_var['reports'].get(), date.today().strftime("%d-%m-%Y"))
            if not os.path.exists(report_folder):
                os.mkdir(report_folder)
            create_results_excel()
    except Exception as e:
        logger.error("Please Select a valid reports folder: " + str(e))


def initialise():
    global report_folder
    global results_xlsx
    try:
        prev_state = None
        if os.path.exists('store.pckl'):
            prev_state = pickle.load(open('store.pckl', 'rb'))
        else:
            logger.warning("Previous state not found. Please intialise new values")
        if prev_state:
            dir_var['xmls'].set(prev_state['xmls_dir'])
            dir_var['reports'].set(prev_state['reports_dir'])
            _set_reports_dir()
        opt.set(1)
        xmls_count_var.set("Select a xml file")
    except Exception as e:
        logger.error(e)


def set_xml_dir():
    folder = askdirectory(title='select xml files folder')
    if folder:
        dir_var['xmls'].set(folder)


def set_reports_dir():
    folder = askdirectory(title='Save report to')
    dir_var['reports'].set(folder)
    _set_reports_dir()


def set_mode(root):
    pass
    global excel_frame
    global xml_df
    file_var['old_xml'].set("")
    file_var['new_xml'].set("")
    clear_results_frame()
    xml_df = None
    xmls_count_var.set("Select a xml file")
    if opt.get() == 1:
        update_logs("######################### Changing to Manual mode #########################")
        excel_frame.pack_forget()
        old_xml_entry.config(state='normal')
        new_xml_entry.config(state='normal')
        cmp_btn.grid()
        total_cnt_label_1.grid_remove()
        total_cnt_label_2.grid_remove()
        next_btn.grid_forget()
        prev_btn.grid_forget()
        cnt_label.grid_forget()
    else:
        update_logs("######################### Changing to Excel mode  #########################")
        excel_frame.pack()
        old_xml_entry.config(state='disabled')
        new_xml_entry.config(state='disabled')
        cmp_btn.grid_remove()
        total_cnt_label_1.grid(row=0, column=4)
        total_cnt_label_2.grid(row=0, column=5)
        prev_btn.grid(row=3, column=0, padx=10, pady=10, columnspan=2)
        cnt_label.grid(row=3, column=2, padx=10)
        next_btn.grid(row=3, column=3, padx=10, columnspan=2)


def get_data_from_excel():
    logger.debug(file_var['excel'].get())
    try:
        df = pd.read_excel(file_var['excel'].get())
        logger.debug(df)
        print(df)
        return df
    except Exception as e:
        logger.error(e)


def set_exl_file():
    global xml_df
    global cur_pos
    initial_dir = os.path.dirname(file_var['excel'].get())
    file = askopenfilename(title="Select Excel", initialdir=initial_dir)
    if file:
        if not file.endswith(".xlsx") and not file.endswith('.csv'):
            msg_box.showerror("Excel file", "Please select a excel file")
            return
        file_var['excel'].set(file)
        xml_df = get_data_from_excel()
        cur_pos = 0
        file_var['old_xml'].set(xml_df.loc[cur_pos]['old'])
        file_var['new_xml'].set(xml_df.loc[cur_pos]['new'])
        xmls_count_var.set(len(xml_df))
        cur_xml_count_var.set(1)
        cmd()
    return


def get_xml_files(folder, file):
    file = "{:0>8}".format(file)
    xml_files = [xml_file for xml_file in os.listdir(folder) if
                 os.path.isfile(os.path.join(folder, xml_file)) and xml_file.startswith(file)]
    return xml_files


def compare(dir, xml1, xml2):
    file1 = os.path.join(dir, xml1)
    file2 = os.path.join(dir, xml2)
    diff = main.diff_files(file1, file2)
    if diff:
        out_file = os.path.join(report_folder, xml1.split('.')[0] + '_' + xml2.split('.')[0] + '.html')
        winmerge = '"%s\WinMerge\WinMergeU.exe"' % os.environ['ProgramFiles']
        cmd_line = '%s /e /ul /ur "%s" "%s" -minimize -noninteractive -u -or "%s"' % (winmerge, file1, file2, out_file)
        sp = subprocess.Popen(cmd_line, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        return "Not identical", out_file
    else:
        return "identical", None


def create_res_frame(xml1, xml2, res, outfile):
    new_res_frame = tk.Frame(results_frame)
    tk.Label(new_res_frame, text=xml1, bd=2, relief="solid", padx=10, pady=10).grid(row=0, column=0)
    tk.Label(new_res_frame, text=xml2, bd=2, relief="solid", padx=10, pady=10).grid(row=0, column=1)
    if res == "identical":
        tk.Label(new_res_frame, text=res, bd=2, bg="green", padx=22, pady=10, relief="sunken")\
            .grid(row=0, column=2,padx=10)
    elif res == "Error":
        tk.Label(new_res_frame, text=res, bd=2, bg="blue", padx=22, pady=10, relief="sunken") \
            .grid(row=0, column=2, padx=10)
    else:
        tk.Button(new_res_frame, text=res, bg="red", padx=8, pady=8,
                  command=lambda: webbrowser.open(outfile, new=2)).grid(row=0, column=2, padx=10)
    new_res_frame.pack()


def save_results(xml1, xml2, res):
    wb = load_workbook(results_xlsx)
    ws = wb.active
    ws.append([datetime.now().time(), xml1, xml2, res])
    if not res == "identical":
        ws[ws.max_row][ws.max_column - 1].font = Font(color=colors.RED)
    wb.save(results_xlsx)
    return

def _compare(xml1, xml2, results):
    start = datetime.now()
    try:
        res, outfile = compare(dir_var['xmls'].get(), xml1, xml2)
        logger.info("{} & {} compared: {} in {}ms".format(xml1, xml2, res,
                                                          (datetime.now() - start) / timedelta(milliseconds=1)))
    except Exception as e:
        res = "Error"
        outfile = None
        logger.error("Error is processing {} & {} {}".format(xml1, xml2, e))
    finally:
        create_res_frame(xml1, xml2, res, outfile)
        # save_results(xml1, xml2, res)
        results.append([xml1, xml2, res])
        update_logs("{} & {} compared: {} in {}ms".format(xml1, xml2, res,
                                                          (datetime.now() - start) / timedelta(milliseconds=1)))

def cmd():
    clear_results_frame()
    file1 = file_var['old_xml'].get()
    file2 = file_var['new_xml'].get()
    if file1 == "" or file2 == "":
        msg_box.showwarning("XML file", "Enter file names")
        return
    xml_files_1 = get_xml_files(dir_var['xmls'].get(), file1)
    xml_files_2 = get_xml_files(dir_var['xmls'].get(), file2)
    logger.debug("old_xmls: {}\n new_xmls: {}".format(xml_files_1, xml_files_2))
    if not xml_files_1 and not xml_files_2:
        msg_box.showerror("XML files", "XML files starting with\n" + file1 + " not found\n" + file2 + " not found")
        return
    elif not xml_files_1:
        msg_box.showerror("XML files", "XML files starting with\n" + file1 + " not found")
        return
    elif not xml_files_2:
        msg_box.showerror("XML files", "XML files starting withn\n" + file2 + " not found")
        return
    if not len(xml_files_1) == len(xml_files_2):
        msg_box.showerror("XML files", "Sub files length mismatch")
        return
    else:
        update_logs("********** Starting to compare {} {} with line items:{}  **********".format(file1, file2, len(xml_files_1)))
        results = []
        for xml1, xml2 in zip(xml_files_1, xml_files_2):
            thread = threading.Thread(target=_compare, args=(xml1, xml2, results, ))
            thread.start()



def on_closing():
    # if msg_box.askokcancel("Quit", "Do you want to quit?"):
    with open('store.pckl', 'wb') as fp:
        current_state = {'xmls_dir': dir_var['xmls'].get(),
                         'reports_dir': dir_var['reports'].get(), }
        pickle.dump(current_state, fp)
    root.destroy()


def show_next(pos):
    global cur_pos
    if xml_df is not None and not xml_df.empty:
        if (pos == 1 and cur_pos < len(xml_df) - 1) or (pos == -1 and cur_pos > 0):
            cur_pos += pos
            file_var['old_xml'].set(xml_df.loc[cur_pos]['old'])
            file_var['new_xml'].set(xml_df.loc[cur_pos]['new'])
            cur_xml_count_var.set(cur_pos + 1)
            cmd()


def clear_results_frame():
    # progress.set(0)
    for widget in results_frame.winfo_children():
        widget.destroy()


def update_logs(data):
    global logs
    global log_var
    logs.append(data)
    if len(logs) > log_cap:
        logs = logs[len(logs) - log_cap - 1:]
    log_var.set("\n".join(logs))


if __name__ == '__main__':
    logger = logging.getLogger("application")
    log_level = logging.INFO
    formatter = logging.Formatter('%(asctime)s [%(levelname)s] %(message)s', "%Y-%m-%d %H:%M:%S")
    handler = logging.StreamHandler(stream=sys.stdout)
    handler.setFormatter(formatter)
    handler.setLevel(log_level)

    logger.addHandler(handler)
    logger.setLevel(log_level)

    root = tk.Tk()
    root.title("XML Compare")
    # root.geometry("640x720")
    root.geometry("{}x{}+0+0".format(root.winfo_screenwidth(), root.winfo_screenheight()))

    dir_var = {"xmls": tk.StringVar(),
               "reports": tk.StringVar()}
    file_var = {"old_xml": tk.StringVar(),
                "new_xml": tk.StringVar(),
                "excel": tk.StringVar()}

    opt = tk.IntVar()
    xmls_count_var = tk.IntVar()
    report_folder = ""
    results_xlsx = ""
    initialise()

    dir_frame = tk.Frame(root)
    dir_frame.pack()

    # XML directory Selection Frame and buttons
    xml_dir_btn = tk.Button(dir_frame, text="Select XML folder", command=set_xml_dir, padx=10)
    xml_dir_btn.grid(row=0, column=0, pady=10, padx=10)
    tk.Label(dir_frame, text="XML folder: ").grid(row=0, column=1, padx=5)
    tk.Label(dir_frame, textvariable=dir_var['xmls'], fg="blue").grid(row=0, column=2)

    # Reports directory Selection Frame
    report_dir_btn = tk.Button(dir_frame, text="Save report to", command=set_reports_dir, padx=20)
    report_dir_btn.grid(row=1, column=0, pady=10, padx=10)
    tk.Label(dir_frame, text="Report folder: ").grid(row=1, column=1, padx=5)
    tk.Label(dir_frame, textvariable=dir_var['reports'], fg="blue").grid(row=1, column=2)

    # # Radio buttons to change b/w manual and excel
    tk.Radiobutton(dir_frame, text="Manual", variable=opt, value=1, command=lambda: set_mode(root))\
        .grid(row=2, column=0, columnspan=2)
    tk.Radiobutton(dir_frame, text="Excel", variable=opt, value=2, command=lambda: set_mode(root))\
        .grid(row=2, column=1, columnspan=2)

    mode_frame = tk.Frame(root)
    mode_frame.pack()
    excel_frame = tk.Frame(mode_frame)

    xml_df = None
    cur_pos = 0
    # Button to select excel file
    exl_file_btn = tk.Button(excel_frame, text="Excel File", command=set_exl_file, padx=20)
    exl_file_btn.grid(row=0, column=0, pady=10, padx=10)
    tk.Label(excel_frame, text="Excel file:").grid(row=0, column=1, padx=5)
    tk.Label(excel_frame, textvariable=file_var['excel'], fg="blue").grid(row=0, column=2)

    # Text Boxes to enter old and new xml line numbers
    entry_frame = tk.Frame(root)
    entry_frame.pack()
    old_xml_entry = tk.Entry(entry_frame, textvariable=file_var['old_xml'])
    new_xml_entry = tk.Entry(entry_frame, textvariable=file_var['new_xml'])

    # Displaying textboxes
    tk.Label(entry_frame, text='Old XMLs:', padx=5).grid(row=0, column=0)
    old_xml_entry.grid(row=0, column=1, padx=10)
    tk.Label(entry_frame, text='New XMLs:', padx=5).grid(row=0, column=2)
    new_xml_entry.grid(row=0, column=3, padx=10)

    # Compare button to start comaprisions
    cmp_btn = tk.Button(entry_frame, text="Compare", command=cmd, padx=10)
    cmp_btn.grid(row=2, column=0, columnspan=4, pady=10)
    root.bind("<Return>", lambda e: cmd() if opt.get() == 1 else None)

    total_cnt_label_1 = tk.Label(entry_frame, text="Count:")
    total_cnt_label_2 = tk.Label(entry_frame, textvariable=xmls_count_var, fg="green")

    # Buttons to select next and prev xmls
    prev_btn = tk.Button(entry_frame, text="prev", command=lambda: show_next(-1), padx=10)
    cur_xml_count_var = tk.IntVar()
    cnt_label = tk.Label(entry_frame, textvariable=cur_xml_count_var)
    next_btn = tk.Button(entry_frame, text="Next", command=lambda: show_next(1), padx=10)

    # Frames to display results
    sf = Pmw.ScrolledFrame(root)
    sf.pack(padx=10, pady=10, ipadx=60, fill='y', expand=1)
    results_frame = sf.interior()

    sf_logging_frame = ScrollableFrame(root, bg="black")
    sf_logging_frame.pack(fill="both", padx=10, pady=10)
    logging_frame = sf_logging_frame.scrollable_frame

    log_var = tk.StringVar()
    logs = []
    log_cap = 400
    tk.Label(logging_frame, textvariable=log_var, bg="black", fg="white", anchor="w").pack(ipadx=640, expand=True)


    root.protocol("WM_DELETE_WINDOW", on_closing)
    root.mainloop()
